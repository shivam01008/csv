# import os
# import pandas as pd
# import requests
# from flask import Flask, render_template, request, send_file
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# from openpyxl.utils import get_column_letter
# from io import BytesIO

# app = Flask(__name__)

# # Config
# UPLOAD_FOLDER = "uploads"
# OUTPUT_FOLDER = "outputs"
# app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# def detect_image_columns(df):
#     image_columns = []

#     for col in df.columns:
#         sample_values = df[col].dropna().astype(str).head(5)

#         for val in sample_values:
#             if val.startswith("http") and (
#                 ".jpg" in val.lower() or
#                 ".jpeg" in val.lower() or
#                 ".png" in val.lower()
#             ):
#                 image_columns.append(col)
#                 break

#     return image_columns


# def create_excel_from_csv(csv_path, output_path):
#     try:
#         df = pd.read_csv(
#             csv_path,
#             engine="python",
#             on_bad_lines="skip"
#         )
#     except Exception as e:
#         print("CSV Error:", e)
#         return False

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Data"

#     # 🔥 AUTO DETECT IMAGE COLUMNS
#     image_columns = detect_image_columns(df)

#     print("Detected Image Columns:", image_columns)

#     all_columns = list(df.columns)

#     # Write headers
#     ws.append(all_columns)

#     # Set column width
#     for i in range(len(all_columns)):
#         col_letter = get_column_letter(i + 1)
#         ws.column_dimensions[col_letter].width = 25

#     row_num = 2

#     for _, row in df.iterrows():
#         col_index = 1

#         for col_name in all_columns:
#             value = row.get(col_name, "")

#             # TEXT FIELD
#             if col_name not in image_columns:
#                 ws.cell(row=row_num, column=col_index, value=value)

#             # IMAGE FIELD
#             else:
#                 if value and isinstance(value, str):

#                     # Skip PDF
#                     if value.lower().endswith(".pdf"):
#                         ws.cell(row=row_num, column=col_index, value="PDF")

#                     else:
#                         try:
#                             response = requests.get(
#                                 value,
#                                 timeout=10,
#                                 headers={"User-Agent": "Mozilla/5.0"}
#                             )

#                             content_type = response.headers.get("Content-Type", "")

#                             print("URL:", value)
#                             print("Type:", content_type)

#                             if response.status_code == 200 and "image" in content_type:
#                                 img_data = BytesIO(response.content)
#                                 img = Image(img_data)

#                                 img.width = 80
#                                 img.height = 80

#                                 cell = f"{get_column_letter(col_index)}{row_num}"
#                                 ws.add_image(img, cell)

#                                 ws.row_dimensions[row_num].height = 65
#                             else:
#                                 ws.cell(row=row_num, column=col_index, value="Not Image")

#                         except Exception as e:
#                             print("Image Error:", e)
#                             ws.cell(row=row_num, column=col_index, value="Error")

#             col_index += 1

#         row_num += 1

#     wb.save(output_path)
#     return True


# @app.route("/", methods=["GET", "POST"])
# def index():
#     if request.method == "POST":
#         try:
#             file = request.files["file"]

#             if file.filename == "":
#                 return "No file selected"

#             upload_path = os.path.join(UPLOAD_FOLDER, file.filename)
#             file.save(upload_path)

#             output_path = os.path.join(OUTPUT_FOLDER, "output.xlsx")

#             success = create_excel_from_csv(upload_path, output_path)

#             if not success:
#                 return "Error processing CSV"

#             return send_file(output_path, as_attachment=True)

#         except Exception as e:
#             return f"Server Error: {str(e)}"

#     return render_template("index.html")


# if __name__ == "__main__":
#     app.run(debug=True)
#................
import os
import pandas as pd
import requests
import time
from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image as PILImage

app = Flask(__name__)

# 🔥 Limit upload size (50MB)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# 🔍 Detect image columns automatically
def detect_image_columns(df):
    image_columns = []

    for col in df.columns:
        sample_values = df[col].dropna().astype(str).head(5)

        for val in sample_values:
            if val.startswith("http"):
                image_columns.append(col)
                break

    return image_columns


# 🔥 Fetch image with retry
def fetch_image_strict(value):
    if not value or not isinstance(value, str):
        return None

    if value.lower().endswith(".pdf"):
        return "PDF"

    for attempt in range(3):
        try:
            response = requests.get(
                value,
                timeout=8,
                headers={"User-Agent": "Mozilla/5.0"}
            )

            if response.status_code == 200:
                return BytesIO(response.content)

        except:
            time.sleep(1)

    return "FAILED"


def create_excel_from_csv(csv_path, output_path):
    try:
        df = pd.read_csv(csv_path, engine="python", on_bad_lines="skip")
    except Exception as e:
        print("CSV Error:", e)
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    image_columns = detect_image_columns(df)
    print("Detected Image Columns:", image_columns)

    all_columns = list(df.columns)
    ws.append(all_columns)

    # Set column width
    for i in range(len(all_columns)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 20

    row_num = 2

    for _, row in df.iterrows():
        col_index = 1

        for col_name in all_columns:
            value = row.get(col_name, "")

            if col_name not in image_columns:
                ws.cell(row=row_num, column=col_index, value=value)

            else:
                result = fetch_image_strict(value)

                if isinstance(result, BytesIO):
                    try:
                        pil_img = PILImage.open(result)

                        if pil_img.mode != "RGB":
                            pil_img = pil_img.convert("RGB")

                        pil_img.thumbnail((200, 200))

                        temp_stream = BytesIO()
                        pil_img.save(temp_stream, format="JPEG", quality=40)
                        temp_stream.seek(0)

                        img = Image(temp_stream)
                        img.width = 50
                        img.height = 50

                        cell = f"{get_column_letter(col_index)}{row_num}"
                        ws.add_image(img, cell)
                        ws.row_dimensions[row_num].height = 40

                    except Exception as e:
                        print("Image Error:", e)
                        ws.cell(row=row_num, column=col_index, value="No Image")

                elif result == "PDF":
                    ws.cell(row=row_num, column=col_index, value="PDF")

                else:
                    ws.cell(row=row_num, column=col_index, value="No Image")

            col_index += 1

        row_num += 1

    wb.save(output_path)
    return True


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            file = request.files.get("file")

            if not file or file.filename == "":
                return "No file selected"

            # Save upload
            upload_path = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(upload_path)

            # Unique output file
            output_path = os.path.join(
                OUTPUT_FOLDER,
                f"output_{int(time.time())}.xlsx"
            )

            success = create_excel_from_csv(upload_path, output_path)

            if not success:
                return "Error processing CSV"

            return send_file(output_path, as_attachment=True)

        except Exception as e:
            return f"Server Error: {str(e)}"

    return render_template("index.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 3000))
    app.run(host="0.0.0.0", port=port)
# .........
# import os
# import pandas as pd
# import requests
# import time
# from flask import Flask, render_template, request, send_file
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# from openpyxl.utils import get_column_letter
# from io import BytesIO
# from PIL import Image as PILImage

# import cv2
# import pytesseract
# import numpy as np

# app = Flask(__name__)

# UPLOAD_FOLDER = "uploads"
# OUTPUT_FOLDER = "outputs"

# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# # ✅ Tesseract path (Mac)
# pytesseract.pytesseract.tesseract_cmd = "/opt/homebrew/bin/tesseract"


# # ---------------- IMAGE QUALITY CHECK ---------------- #

# def check_image_quality(image_path):
#     img = cv2.imread(image_path)

#     if img is None:
#         return False, "Image not readable"

#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

#     blur = cv2.Laplacian(gray, cv2.CV_64F).var()
#     brightness = np.mean(gray)
#     h, w = gray.shape

#     print(f"Blur={blur}, Brightness={brightness}, Size={w}x{h}")

#     if blur < 50:
#         return False, "Blurry image"

#     if brightness < 50 or brightness > 220:
#         return False, "Lighting issue"

#     if w < 800 or h < 400:
#         return False, "Low resolution"

#     return True, "OK"


# # ---------------- CSV → EXCEL ---------------- #

# def detect_image_columns(df):
#     image_columns = []

#     for col in df.columns:
#         sample = df[col].dropna().astype(str).head(5)
#         for val in sample:
#             if val.startswith("http"):
#                 image_columns.append(col)
#                 break

#     return image_columns


# def fetch_image_strict(value):
#     if not value or not isinstance(value, str):
#         return None

#     if value.lower().endswith(".pdf"):
#         return "PDF"

#     for _ in range(5):
#         try:
#             r = requests.get(value, timeout=15)
#             if r.status_code == 200:
#                 return BytesIO(r.content)
#         except:
#             time.sleep(2)

#     return "FAILED"


# def create_excel_from_csv(csv_path, output_path):
#     try:
#         df = pd.read_csv(csv_path, engine="python", on_bad_lines="skip")
#     except Exception as e:
#         print("CSV Error:", e)
#         return False

#     wb = Workbook()
#     ws = wb.active

#     cols = list(df.columns)
#     img_cols = detect_image_columns(df)

#     ws.append(cols)

#     for i in range(len(cols)):
#         ws.column_dimensions[get_column_letter(i+1)].width = 20

#     r = 2

#     for _, row in df.iterrows():
#         c = 1

#         for col in cols:
#             val = row.get(col, "")

#             if col not in img_cols:
#                 ws.cell(row=r, column=c, value=val)

#             else:
#                 img_data = fetch_image_strict(val)

#                 if isinstance(img_data, BytesIO):
#                     try:
#                         img = PILImage.open(img_data)
#                         img.thumbnail((200, 200))

#                         tmp = BytesIO()
#                         img.save(tmp, format="JPEG", quality=40)
#                         tmp.seek(0)

#                         ws.add_image(Image(tmp), f"{get_column_letter(c)}{r}")
#                     except:
#                         ws.cell(row=r, column=c, value="Image Error")

#                 else:
#                     ws.cell(row=r, column=c, value="Failed")

#             c += 1

#         r += 1

#     wb.save(output_path)
#     return True


# # ---------------- IMAGE → EXCEL (STRUCTURED) ---------------- #

# def create_excel_from_image(image_path, output_path):
#     try:
#         valid, msg = check_image_quality(image_path)
#         if not valid:
#             print("❌", msg)
#             return False

#         img = cv2.imread(image_path)
#         gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

#         _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

#         horiz = cv2.morphologyEx(
#             thresh, cv2.MORPH_OPEN,
#             cv2.getStructuringElement(cv2.MORPH_RECT, (60, 1))
#         )

#         vert = cv2.morphologyEx(
#             thresh, cv2.MORPH_OPEN,
#             cv2.getStructuringElement(cv2.MORPH_RECT, (1, 60))
#         )

#         mask = cv2.add(horiz, vert)

#         contours, _ = cv2.findContours(mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

#         boxes = []
#         for cnt in contours:
#             x, y, w, h = cv2.boundingRect(cnt)
#             if w < 80 or h < 25:
#                 continue
#             boxes.append((x, y, w, h))

#         boxes = sorted(boxes, key=lambda b: (b[1], b[0]))

#         rows = []
#         current = []
#         last_y = -100

#         for (x, y, w, h) in boxes:
#             if abs(y - last_y) > 20:
#                 if current:
#                     rows.append(current)
#                 current = [(x, y, w, h)]
#                 last_y = y
#             else:
#                 current.append((x, y, w, h))

#         if current:
#             rows.append(current)

#         data = []

#         for row in rows:
#             row = sorted(row, key=lambda b: b[0])
#             row_data = []

#             for (x, y, w, h) in row:
#                 cell = gray[y:y+h, x:x+w]
#                 text = pytesseract.image_to_string(cell, config="--psm 7").strip()
#                 row_data.append(text)

#             data.append(row_data)

#         # ✅ FIX: STRUCTURED OUTPUT
#         columns = [
#             "CLAIM NO",
#             "INSURED",
#             "LOSS_DATE",
#             "VEHICLE_REG_NO",
#             "HUB",
#             "VEHICLE_MAKE",
#             "INSURED TEL"
#         ]

#         clean_rows = []

#         for row in data:
#             if len(row) < len(columns):
#                 row += [""] * (len(columns) - len(row))
#             else:
#                 row = row[:len(columns)]

#             clean_rows.append(row)

#         df = pd.DataFrame(clean_rows, columns=columns)

#         # Remove header if duplicated
#         if "CLAIM" in str(df.iloc[0][0]).upper():
#             df = df[1:].reset_index(drop=True)

#         # Clean whitespace
#         df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

#         df.to_excel(output_path, index=False)

#         return True

#     except Exception as e:
#         print("Image Error:", e)
#         return False


# # ---------------- ROUTE ---------------- #

# @app.route("/", methods=["GET", "POST"])
# def index():
#     if request.method == "POST":
#         try:
#             file = request.files["file"]

#             if file.filename == "":
#                 return "No file selected"

#             path = os.path.join(UPLOAD_FOLDER, file.filename)
#             file.save(path)

#             out = os.path.join(OUTPUT_FOLDER, "output.xlsx")

#             name = file.filename.lower()

#             if name.endswith(".csv"):
#                 success = create_excel_from_csv(path, out)

#             elif name.endswith((".jpg", ".png", ".jpeg")):
#                 success = create_excel_from_image(path, out)

#             else:
#                 return "Unsupported file type"

#             if not success:
#                 return "Image quality not good or processing failed"

#             return send_file(out, as_attachment=True)

#         except Exception as e:
#             import traceback
#             traceback.print_exc()
#             return f"Error: {str(e)}"

#     return render_template("index.html")


# if __name__ == "__main__":
#     app.run(debug=True)